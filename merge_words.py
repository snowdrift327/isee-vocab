"""
Merge two word libraries into one deduplicated master file.

Usage:
  uv run merge_words.py
"""

import openpyxl
from pathlib import Path
from datetime import datetime

# ============ Configuration ============
OLD_FILE = Path("isee_lower_level_words.xlsx")
NEW_FILE = Path("isee_lower_level_words_2.xlsx")
OUTPUT_FILE = Path("isee_lower_level_words_v3.xlsx")


def read_words_from_excel(filepath):
    """Read all words from all sheets and all columns."""
    print(f"\n📖 Reading {filepath}...")
    if not filepath.exists():
        print(f"   ⚠️  File not found: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath)
    words = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_count = 0
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                word = str(cell).strip()
                # Only alphabetic words with 3+ characters
                if word and word.isalpha() and len(word) >= 3:
                    words.append(word)
                    sheet_count += 1
        print(f"   Sheet [{sheet_name}]: {sheet_count} words")

    print(f"   Total (with duplicates within file): {len(words)}")
    return words


def normalize(word):
    """Lowercase for comparison, but preserve original casing."""
    return word.lower()


def main():
    # Read both files
    old_words = read_words_from_excel(OLD_FILE)
    new_words = read_words_from_excel(NEW_FILE)

    # Merge and dedup (case-insensitive)
    all_words = old_words + new_words
    seen = set()
    unique_words = []

    # Track statistics
    from_old_only = 0
    from_new_only = 0
    duplicates_between = 0

    old_set = {normalize(w) for w in old_words}
    new_set = {normalize(w) for w in new_words}

    overlap = old_set & new_set
    only_old = old_set - new_set
    only_new = new_set - old_set

    # Build final list (capitalize each word)
    for w in all_words:
        norm = normalize(w)
        if norm not in seen:
            seen.add(norm)
            unique_words.append(w.capitalize())

    # Sort alphabetically for clean output
    unique_words.sort()

    # Statistics
    print(f"\n{'=' * 60}")
    print(f"📊 MERGE STATISTICS")
    print(f"{'=' * 60}")
    print(f"Words in old file:               {len(old_set)} unique")
    print(f"Words in new file:               {len(new_set)} unique")
    print(f"Words in BOTH files (overlap):   {len(overlap)}")
    print(f"Words ONLY in old file:          {len(only_old)}")
    print(f"Words ONLY in new file:          {len(only_new)}")
    print(f"─" * 60)
    print(f"Final merged (deduplicated):     {len(unique_words)} unique words")
    print(f"{'=' * 60}\n")

    # Write output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged Vocabulary"
    for word in unique_words:
        ws_out.append([word])
    wb_out.save(OUTPUT_FILE)

    print(f"✅ Merged file saved to: {OUTPUT_FILE}")
    print(f"   Total: {len(unique_words)} unique words")

    # Estimate quiz set capacity
    max_sets = len(unique_words) // 20
    print(f"\n💡 With 20 words per set, you can generate {max_sets} unique quiz sets.")


if __name__ == "__main__":
    main()