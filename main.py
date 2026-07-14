"""
ISEE Vocabulary — Batch Generator
Generates N quiz sets at once, saves to quiz-sets.json, supports resume on interruption.

Usage:
  uv run main.py            → generate 30 quiz sets (default)
  uv run main.py 20         → generate 20 quiz sets
  uv run main.py reset      → clear quiz-sets.json and start over
"""

import os
import json
import random
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from anthropic import Anthropic
from dotenv import load_dotenv

# ============ Configuration ============
load_dotenv()
client = Anthropic()

WORDS_FILE = Path("isee_lower_level_words.xlsx")
QUIZ_SETS_FILE = Path("quiz-sets.json")
INDEX_FILE = Path("index.html")
HISTORY_PAGE = Path("history.html")
MISTAKES_FILE = Path("mistakes.html")

DEFAULT_BATCH_SIZE = 30
QUESTIONS_PER_SET = 20
SYNONYM_COUNT = 14  # 14 synonym questions
SENTENCE_COUNT = 6  # 6 sentence completion questions


# ============ Load Words ============
def load_all_words():
    """Read ALL words from ALL sheets and ALL columns (skip empty cells)."""
    wb = openpyxl.load_workbook(WORDS_FILE)
    words = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                word = str(cell).strip()
                # Filter: must be a real word (alpha only, 2+ chars)
                # Skip headers, numbers, and non-word content
                if word and word.isalpha() and len(word) >= 3:
                    words.append(word.lower())  # normalize to lowercase for dedup
    # Dedup and restore case (first-seen wins)
    seen = set()
    unique = []
    for w in words:
        if w not in seen:
            seen.add(w)
            # Capitalize first letter to match your Excel format
            unique.append(w.capitalize())
    return unique


# ============ AI Generation ============
def parse_ai_json(prompt, max_tokens=4096, max_retries=3, label=""):
    """Call Claude and parse JSON, retry on failure."""
    last_error = None
    for attempt in range(max_retries):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=max_tokens,
                messages=[{"role": "user", "content": prompt}]
            )
            text = response.content[0].text.strip()
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
                text = text.strip()
            return json.loads(text)
        except json.JSONDecodeError as e:
            last_error = e
            print(f"   ⚠️  {label} attempt {attempt+1}/{max_retries}: JSON parse failed (line {e.lineno}, col {e.colno})")
            if attempt < max_retries - 1:
                print(f"   🔄 Retrying...")
            continue
    raise RuntimeError(f"{label} failed after {max_retries} attempts. Last error: {last_error}")


def generate_synonym_questions(words_batch):
    """Generate synonym questions for a batch of words."""
    words_str = ", ".join(words_batch)
    prompt = f"""You are a test question writer for the ISEE Lower Level exam (grades 5-6 admission). Create {len(words_batch)} synonym multiple-choice questions.

WORDS TO USE (one question per word): {words_str}

FORMAT for each question:
- Stem: "[WORD] most nearly means"
- 4 options: 1 correct synonym + 3 plausible distractors
- Distractors MUST be from the SAME PART OF SPEECH as the correct answer
- Distractors should be words the student might confuse (similar sound, related meaning, common wrong answer)
- Difficulty must match ISEE Lower Level — challenging but fair for advanced 5-6th graders
- Include a 1-sentence explanation

DISTRACTOR QUALITY RULES:
- Distractor 1: Similar meaning but wrong nuance (partially right)
- Distractor 2: Related concept but different meaning
- Distractor 3: Common wrong answer students give

Output STRICT JSON only:
{{
  "questions": [
    {{
      "type": "synonym",
      "word": "WORD",
      "stem": "WORD most nearly means",
      "options": ["option1", "option2", "option3", "option4"],
      "correct_index": 0,
      "explanation": "..."
    }},
    ...
  ]
}}

Generate exactly {len(words_batch)} questions."""

    result = parse_ai_json(prompt, max_tokens=6144, label=f"Synonym batch ({len(words_batch)})")
    return result["questions"]


def generate_sentence_questions(words_batch):
    """Generate sentence completion questions for a batch of words."""
    words_str = ", ".join(words_batch)
    prompt = f"""You are a test question writer for the ISEE Lower Level exam. Create {len(words_batch)} sentence completion questions.

WORDS TO USE (one question per word): {words_str}

FORMAT for each question:
- A single sentence with ONE blank ______
- The sentence context clearly points to the target word
- 4 options: 1 correct + 3 plausible distractors
- All 4 options must be the SAME PART OF SPEECH
- Distractors must fit grammatically but not semantically
- Include a 1-sentence explanation

QUALITY RULES:
- Sentence should be 12-25 words long
- Context clues should be specific (a cause, a contrast, an example)
- Avoid trivial fillings — the correct word should be clearly best but require thinking
- Difficulty matches ISEE Lower Level

Output STRICT JSON only:
{{
  "questions": [
    {{
      "type": "sentence",
      "word": "WORD",
      "stem": "The sentence with a ______ blank in it.",
      "options": ["option1", "option2", "option3", "option4"],
      "correct_index": 0,
      "explanation": "..."
    }},
    ...
  ]
}}

Generate exactly {len(words_batch)} questions."""

    result = parse_ai_json(prompt, max_tokens=6144, label=f"Sentence batch ({len(words_batch)})")
    return result["questions"]


def generate_one_quiz_set(set_number, chosen_words):
    """Generate one complete quiz set (20 questions)."""
    synonym_words = chosen_words[:SYNONYM_COUNT]
    sentence_words = chosen_words[SYNONYM_COUNT:SYNONYM_COUNT + SENTENCE_COUNT]

    print(f"   Generating {SYNONYM_COUNT} synonym questions...")
    synonym_qs = generate_synonym_questions(synonym_words)

    print(f"   Generating {SENTENCE_COUNT} sentence completion questions...")
    sentence_qs = generate_sentence_questions(sentence_words)

    # Interleave: alternate synonym/sentence for variety, but keep majority synonym
    all_questions = synonym_qs + sentence_qs
    random.shuffle(all_questions)

    return {
        "id": f"set_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
        "set_number": set_number,
        "created_at": datetime.now().isoformat(),
        "words_used": chosen_words,
        "questions": all_questions
    }


# ============ Load/Save Quiz Sets Database ============
def load_quiz_sets():
    if QUIZ_SETS_FILE.exists():
        with open(QUIZ_SETS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"generated_at": None, "total": 0, "sets": []}


def save_quiz_sets(data):
    data["total"] = len(data["sets"])
    data["last_updated"] = datetime.now().isoformat()
    with open(QUIZ_SETS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ============ Main Batch Generation ============
def run_batch(target_count):
    all_words = load_all_words()
    print(f"📚 Loaded {len(all_words)} unique words from {WORDS_FILE}")

    db = load_quiz_sets()
    existing_sets = db.get("sets", [])
    already_have = len(existing_sets)

    to_generate = target_count - already_have
    if to_generate <= 0:
        print(f"✅ Already have {already_have} quiz sets. To start over, run: uv run main.py reset")
        return existing_sets

    # Track used words across all existing sets to avoid overlap
    used_words = set()
    for s in existing_sets:
        used_words.update(w.lower() for w in s.get("words_used", []))

    words_per_set = QUESTIONS_PER_SET
    available_words = [w for w in all_words if w.lower() not in used_words]

    if len(available_words) < to_generate * words_per_set:
        max_possible_sets = already_have + len(available_words) // words_per_set
        print(f"⚠️  Only {len(available_words)} unused words left; can generate max {max_possible_sets - already_have} more sets.")
        to_generate = max_possible_sets - already_have
        if to_generate <= 0:
            print("Cannot generate more sets — no unused words. Consider reset.")
            return existing_sets

    print(f"\n📝 Generating {to_generate} new quiz sets (currently have {already_have})")
    print(f"⏱  Estimated time: {to_generate * 15 // 60}-{to_generate * 30 // 60} minutes")
    print(f"💰 Estimated cost: ${to_generate * 0.06:.2f}-${to_generate * 0.10:.2f}\n")

    if db.get("generated_at") is None:
        db["generated_at"] = datetime.now().isoformat()

    random.shuffle(available_words)
    failed = []

    for i in range(to_generate):
        set_number = already_have + i + 1
        chosen_words = available_words[:words_per_set]
        available_words = available_words[words_per_set:]

        print(f"[{set_number}/{target_count}] Quiz Set #{set_number}")

        try:
            quiz_set = generate_one_quiz_set(set_number, chosen_words)
            existing_sets.append(quiz_set)
            db["sets"] = existing_sets
            save_quiz_sets(db)
            print(f"    ✓ Generated {len(quiz_set['questions'])} questions\n")
        except Exception as e:
            failed.append((set_number, str(e)))
            print(f"    ❌ Failed: {e}\n")
            # Put words back if failed
            available_words = chosen_words + available_words
            continue

    print(f"\n{'=' * 60}")
    print(f"✅ Batch complete!")
    print(f"   Total quiz sets: {len(existing_sets)}")
    if failed:
        print(f"   Failed: {len(failed)}")
    print(f"{'=' * 60}\n")

    return existing_sets


# ============ HTML Generators ============
def render_index_html():
    return '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ISEE Vocabulary Practice</title>
<style>
* { box-sizing: border-box; }
body {
    font-family: "Georgia", "Times New Roman", serif;
    max-width: 760px;
    margin: 0 auto;
    padding: 24px 22px 80px;
    background: #fafaf7;
    color: #1a1a1a;
    line-height: 1.7;
    font-size: 17px;
}
.start-overlay {
    position: fixed; top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(250, 250, 247, 0.98); z-index: 9999;
    display: flex; align-items: center; justify-content: center;
    backdrop-filter: blur(8px);
}
.start-overlay.hidden { display: none; }
.start-card {
    background: white; max-width: 500px; width: 90%;
    padding: 40px 32px; border: 3px double #1a4d8f;
    border-radius: 8px; text-align: center;
    box-shadow: 0 8px 32px rgba(0,0,0,0.15);
}
.start-card h2 { font-size: 26px; color: #1a4d8f; margin: 0 0 8px; }
.start-card .subtitle { color: #666; font-size: 15px; font-style: italic; margin: 8px 0 20px; }
.start-card .instructions {
    text-align: left; background: #f5f5f0; padding: 16px 20px;
    border-radius: 6px; margin: 20px 0 28px;
    font-family: "Helvetica Neue", Arial, sans-serif;
    font-size: 14px; line-height: 1.7;
}
.start-card .instructions strong {
    color: #1a4d8f; display: block; margin-bottom: 6px;
    font-size: 13px; text-transform: uppercase; letter-spacing: 1px;
}
.start-card .instructions ul { margin: 0; padding-left: 20px; }
.start-btn {
    background: #1a4d8f; color: white; border: none;
    padding: 16px 56px; font-size: 17px; font-weight: bold;
    border-radius: 4px; cursor: pointer;
    font-family: "Helvetica Neue", Arial, sans-serif;
    letter-spacing: 2px;
}
.start-btn:hover { background: #143a6b; }
header { border-bottom: 3px double #333; padding-bottom: 20px; margin-bottom: 28px; }
h1 { font-size: 26px; margin: 0 0 8px; letter-spacing: 1px; }
.subtitle { font-size: 14px; color: #555; font-style: italic; }
.meta-bar {
    display: flex; justify-content: space-between; align-items: center;
    margin-top: 14px; padding: 10px 16px; background: white;
    border: 1px solid #ddd; border-radius: 4px;
    font-family: "Helvetica Neue", Arial, sans-serif; font-size: 14px;
}
.timer { font-family: "Courier New", monospace; font-size: 17px; font-weight: bold; color: #1a4d8f; }
.progress-pill { background: white; padding: 6px 12px; border-radius: 12px; font-size: 12px; color: #555; border: 1px solid #ddd; font-family: Arial, sans-serif; }
.question {
    background: white; padding: 22px 28px; border-radius: 8px;
    margin-bottom: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
}
.q-header { margin-bottom: 12px; font-family: "Helvetica Neue", Arial, sans-serif; }
.q-number { font-size: 18px; font-weight: bold; margin-right: 12px; }
.q-type {
    font-size: 11px; color: #666;
    font-style: italic; letter-spacing: 1px;
    text-transform: uppercase;
}
.q-stem { font-size: 17px; margin-bottom: 14px; line-height: 1.6; }
.q-stem .highlight { font-weight: bold; letter-spacing: 0.5px; }
.q-options { display: flex; flex-direction: column; gap: 6px; margin-left: 28px; font-family: "Georgia", serif; }
.option {
    display: flex; align-items: center; padding: 10px 14px;
    border: 1.5px solid #e0e0e0; border-radius: 5px;
    cursor: pointer; font-size: 15px;
}
.option:hover { border-color: #1a4d8f; background: #f0f5fb; }
.option input { margin-right: 10px; }
.letter { font-weight: bold; color: #555; margin-right: 12px; min-width: 20px; font-style: italic; }
.option.user-correct { background: #e8f5e9; border-color: #4caf50; }
.option.user-wrong { background: #ffebee; border-color: #f44336; }
.option.show-correct { background: #e8f5e9; border-color: #4caf50; border-style: dashed; }
.submit-section { margin-top: 32px; text-align: center; padding-top: 20px; border-top: 3px double #333; }
#submit-btn {
    background: #1a4d8f; color: white; border: none;
    padding: 14px 52px; font-size: 16px; font-weight: bold;
    border-radius: 4px; cursor: pointer;
    font-family: "Helvetica Neue", Arial, sans-serif;
    letter-spacing: 1px;
}
#submit-btn:hover { background: #143a6b; }
#submit-btn:disabled { background: #aaa; cursor: not-allowed; }
#results {
    display: none; margin-top: 30px; padding: 28px;
    background: #f9f9f7; border: 2px solid #1a4d8f;
    border-radius: 8px; font-family: "Helvetica Neue", Arial, sans-serif;
}
#results.show { display: block; }
.new-test-section { text-align: center; margin-bottom: 20px; padding-bottom: 18px; border-bottom: 1px solid #ddd; }
.start-new-btn {
    background: #2e7d32; color: white; border: none;
    padding: 12px 40px; font-size: 15px; font-weight: bold;
    border-radius: 6px; cursor: pointer; letter-spacing: 1px;
}
.start-new-btn:hover { background: #1b5e20; }
.score-summary { display: flex; justify-content: space-around; text-align: center; margin-bottom: 24px; flex-wrap: wrap; gap: 14px; }
.score-item { flex: 1; min-width: 120px; }
.score-value { font-size: 32px; font-weight: bold; color: #1a4d8f; display: block; }
.score-label { font-size: 12px; color: #666; text-transform: uppercase; letter-spacing: 1px; margin-top: 4px; }
.action-links {
    margin: 20px 0; padding: 16px 0;
    border-top: 1px solid #ddd; border-bottom: 1px solid #ddd;
    display: grid; grid-template-columns: 1fr 1fr; gap: 10px;
}
.action-links a {
    display: block; text-align: center; padding: 12px 16px;
    background: #1a4d8f; color: white; text-decoration: none;
    border-radius: 4px; font-size: 14px; font-weight: 500;
}
.action-links a:hover { background: #143a6b; }
.explanation-block { background: white; padding: 14px 18px; margin-bottom: 12px; border-radius: 6px; border-left: 4px solid #4caf50; }
.explanation-block.wrong { border-left-color: #f44336; }
.exp-q { font-weight: 600; margin-bottom: 6px; font-size: 14px; }
.exp-result { font-size: 13px; color: #555; margin-bottom: 8px; }
.exp-text { font-size: 14px; color: #333; line-height: 1.6; }
footer { text-align: center; margin-top: 50px; color: #999; font-size: 13px; font-family: "Helvetica Neue", Arial, sans-serif; }
footer a { color: #888; text-decoration: none; }
.all-done { text-align: center; padding: 60px 20px; }
.all-done h1 { color: #1a4d8f; font-size: 32px; border: none; }
.all-done .actions { margin-top: 32px; display: flex; gap: 14px; justify-content: center; flex-wrap: wrap; }
.all-done button { background: #1a4d8f; color: white; border: none; padding: 14px 28px; font-size: 14px; font-weight: bold; border-radius: 6px; cursor: pointer; font-family: "Helvetica Neue", Arial, sans-serif; }
.all-done button.secondary { background: #6c757d; }
@media (max-width: 600px) {
    body { padding: 18px 14px 60px; font-size: 16px; }
    h1 { font-size: 22px; }
    .question { padding: 18px 16px; }
    .q-options { margin-left: 8px; }
    .meta-bar { flex-direction: column; gap: 6px; align-items: flex-start; }
    .action-links { grid-template-columns: 1fr; }
}
</style>
</head>
<body>

<div id="start-overlay" class="start-overlay">
    <div class="start-card" id="start-card">
        <div style="text-align:center;color:#999;font-size:14px;">Loading quiz library...</div>
    </div>
</div>

<header id="main-header" style="display:none;">
    <h1>ISEE LOWER LEVEL</h1>
    <div class="subtitle">Vocabulary Practice — Synonyms &amp; Sentence Completion</div>
    <div class="meta-bar">
        <span class="timer">Time: <span id="timer-display">00:00</span></span>
        <span>Questions: <span id="progress-count">0</span> / 20</span>
        <span class="progress-pill" id="library-progress">Loading...</span>
    </div>
</header>

<form id="quiz-form" onsubmit="return false;"></form>

<div class="submit-section" id="submit-section" style="display:none;">
    <button id="submit-btn" onclick="submitQuiz()">SUBMIT</button>
</div>

<div id="results"></div>

<footer style="display:none;" id="footer">
    <a href="history.html">📊 View History</a>
    &nbsp;|&nbsp;
    <a href="mistakes.html">📖 Mistakes Book</a>
    &nbsp;|&nbsp;
    <span id="library-count-footer"></span>
</footer>

<script>
let ALL_SETS = [];
let CURRENT = null;
let CURRENT_ID = null;
let startTime = null;
let submitted = false;
let testStarted = false;

async function init() {
    try {
        const res = await fetch('quiz-sets.json?v=' + Date.now());
        const data = await res.json();
        ALL_SETS = data.sets || [];

        if (ALL_SETS.length === 0) {
            showError("No quiz sets in library. Run 'uv run main.py' on your computer.");
            return;
        }

        const savedId = sessionStorage.getItem('current_set_id');
        if (savedId) {
            CURRENT = ALL_SETS.find(s => s.id === savedId);
            CURRENT_ID = savedId;
        }

        if (!CURRENT) {
            const done = JSON.parse(localStorage.getItem('done_set_ids') || '[]');
            const available = ALL_SETS.filter(s => !done.includes(s.id));

            if (available.length === 0) {
                showAllDone();
                return;
            }

            CURRENT = available[Math.floor(Math.random() * available.length)];
            CURRENT_ID = CURRENT.id;
            sessionStorage.setItem('current_set_id', CURRENT_ID);
        }

        renderSet();

        const savedState = sessionStorage.getItem('vocab_submitted_state_' + CURRENT_ID);
        if (savedState) {
            const state = JSON.parse(savedState);
            document.getElementById('start-overlay').classList.add('hidden');
            document.getElementById('quiz-form').innerHTML = state.formHTML;
            document.getElementById('results').innerHTML = state.resultsHTML;
            document.getElementById('results').classList.add('show');
            document.getElementById('timer-display').textContent = state.timer;
            document.getElementById('progress-count').textContent = state.progress;
            document.getElementById('submit-btn').disabled = true;
            document.getElementById('submit-btn').textContent = 'SUBMITTED';
            submitted = true;
        }
    } catch (err) {
        showError("Failed to load quiz sets: " + err.message);
    }
}

function renderSet() {
    const done = JSON.parse(localStorage.getItem('done_set_ids') || '[]');
    const remaining = ALL_SETS.length - done.length;

    document.getElementById('start-card').innerHTML = `
        <h2>ISEE VOCABULARY</h2>
        <div class="subtitle">Quiz Set #${CURRENT.set_number || '?'}</div>
        <div class="instructions">
            <strong>Instructions</strong>
            <ul>
                <li>20 questions: 14 synonyms + 6 sentence completions</li>
                <li>Read each question carefully</li>
                <li>Select the best answer</li>
                <li>Timer starts when you click START</li>
            </ul>
        </div>
        <button class="start-btn" onclick="startTest()">START</button>
        <div style="margin-top:16px;font-size:12px;color:#888;font-family:Arial;">
            📚 ${ALL_SETS.length - remaining}/${ALL_SETS.length} sets completed
        </div>
    `;

    document.getElementById('library-progress').textContent = `${ALL_SETS.length - remaining}/${ALL_SETS.length} done`;
    document.getElementById('library-count-footer').textContent = `Library: ${ALL_SETS.length} sets`;

    let questionsHTML = "";
    CURRENT.questions.forEach((q, i) => {
        const typeLabel = q.type === "synonym" ? "SYNONYM" : "SENTENCE COMPLETION";
        let stemHTML = q.stem;
        if (q.type === "synonym" && q.word) {
            stemHTML = `<span class="highlight">${q.word}</span> most nearly means`;
        }

        let optionsHTML = "";
        const letters = ["A", "B", "C", "D"];
        q.options.forEach((opt, j) => {
            optionsHTML += `
                <label class="option" data-q="${i}" data-opt="${j}">
                    <input type="radio" name="q${i}" value="${j}">
                    <span class="letter">${letters[j]}</span>
                    <span>${opt}</span>
                </label>`;
        });
        questionsHTML += `
            <div class="question" data-q-index="${i}">
                <div class="q-header">
                    <span class="q-number">${i+1}.</span>
                    <span class="q-type">${typeLabel}</span>
                </div>
                <div class="q-stem">${stemHTML}</div>
                <div class="q-options">${optionsHTML}</div>
            </div>`;
    });
    document.getElementById('quiz-form').innerHTML = questionsHTML;

    document.querySelectorAll('input[type="radio"]').forEach(input => {
        input.addEventListener('change', () => {
            const answered = new Set();
            document.querySelectorAll('input[type="radio"]:checked').forEach(r => answered.add(r.name));
            document.getElementById('progress-count').textContent = answered.size;
        });
    });

    document.getElementById('main-header').style.display = 'block';
    document.getElementById('submit-section').style.display = 'block';
    document.getElementById('footer').style.display = 'block';
}

function startTest() {
    testStarted = true;
    startTime = Date.now();
    document.getElementById('start-overlay').classList.add('hidden');
}

setInterval(() => {
    if (!testStarted || submitted) return;
    const elapsed = Math.floor((Date.now() - startTime) / 1000);
    const mm = String(Math.floor(elapsed / 60)).padStart(2, '0');
    const ss = String(elapsed % 60).padStart(2, '0');
    document.getElementById('timer-display').textContent = mm + ':' + ss;
}, 1000);

function submitQuiz() {
    if (submitted) return;
    const userAnswers = {};
    let answeredCount = 0;
    CURRENT.questions.forEach((q, i) => {
        const sel = document.querySelector('input[name="q' + i + '"]:checked');
        if (sel) { userAnswers[i] = parseInt(sel.value); answeredCount++; }
        else { userAnswers[i] = -1; }
    });

    if (answeredCount < CURRENT.questions.length) {
        if (!confirm('You have ' + (CURRENT.questions.length - answeredCount) + ' unanswered. Submit?')) return;
    }

    submitted = true;
    const totalTime = startTime ? Math.floor((Date.now() - startTime) / 1000) : 0;

    let correctCount = 0;
    const mistakes = [];
    CURRENT.questions.forEach((q, i) => {
        const userAns = userAnswers[i];
        const correctAns = q.correct_index;
        const opts = document.querySelectorAll('label.option[data-q="' + i + '"]');
        opts.forEach(opt => {
            const idx = parseInt(opt.dataset.opt);
            opt.style.pointerEvents = 'none';
            opt.querySelector('input').disabled = true;
            if (idx === correctAns) {
                opt.classList.add(userAns === correctAns ? 'user-correct' : 'show-correct');
            } else if (idx === userAns) {
                opt.classList.add('user-wrong');
            }
        });
        if (userAns === correctAns) correctCount++;
        else {
            mistakes.push({
                word: q.word,
                type: q.type,
                stem: q.stem,
                correct: q.options[correctAns],
                user: userAns >= 0 ? q.options[userAns] : '(unanswered)',
                explanation: q.explanation,
                date: new Date().toISOString().split('T')[0]
            });
        }
    });

    const accuracy = Math.round((correctCount / CURRENT.questions.length) * 100);
    const mm = String(Math.floor(totalTime / 60)).padStart(2, '0');
    const ss = String(totalTime % 60).padStart(2, '0');

    const sessionRecord = {
        date: new Date().toISOString().split('T')[0],
        timestamp: new Date().toISOString(),
        set_id: CURRENT.id,
        set_number: CURRENT.set_number,
        correct: correctCount,
        total: CURRENT.questions.length,
        accuracy: accuracy,
        duration_sec: totalTime
    };
    const allSessions = JSON.parse(localStorage.getItem('vocab_sessions') || '[]');
    allSessions.push(sessionRecord);
    localStorage.setItem('vocab_sessions', JSON.stringify(allSessions));

    // Update mistakes book
    if (mistakes.length > 0) {
        const existingMistakes = JSON.parse(localStorage.getItem('vocab_mistakes') || '[]');
        mistakes.forEach(m => {
            // Don't add duplicate if same word already in mistakes
            if (!existingMistakes.some(em => em.word === m.word && em.type === m.type)) {
                existingMistakes.push(m);
            }
        });
        localStorage.setItem('vocab_mistakes', JSON.stringify(existingMistakes));
    }

    let explanationsHtml = '<h3 style="margin-top:20px;color:#1a1a1a;font-size:16px;border-bottom:1px solid #ddd;padding-bottom:8px;">📝 Review</h3>';
    CURRENT.questions.forEach((q, i) => {
        const userAns = userAnswers[i];
        const isCorrect = userAns === q.correct_index;
        const userText = userAns >= 0 ? q.options[userAns] : '(unanswered)';
        explanationsHtml += '<div class="explanation-block' + (isCorrect ? '' : ' wrong') + '">' +
            '<div class="exp-q">' + (i+1) + '. ' + q.stem + '</div>' +
            '<div class="exp-result">' + (isCorrect ? '✓ Correct: ' : '✗ You answered: ') +
            '<strong>' + userText + '</strong>' +
            (isCorrect ? '' : ' &nbsp; Correct: <strong>' + q.options[q.correct_index] + '</strong>') +
            '</div>' +
            '<div class="exp-text"><strong>Why:</strong> ' + q.explanation + '</div>' +
            '</div>';
    });

    const resultsHtml =
        '<div class="new-test-section">' +
        '<button class="start-new-btn" onclick="nextSet()">🔄 Start New Test</button>' +
        '<p style="color:#666;font-size:12px;margin-top:8px;">Marks this set complete, randomly picks next</p>' +
        '</div>' +
        '<div class="score-summary">' +
        '<div class="score-item"><span class="score-value">' + correctCount + '/' + CURRENT.questions.length + '</span><span class="score-label">Score</span></div>' +
        '<div class="score-item"><span class="score-value">' + accuracy + '%</span><span class="score-label">Accuracy</span></div>' +
        '<div class="score-item"><span class="score-value">' + mm + ':' + ss + '</span><span class="score-label">Time</span></div>' +
        '</div>' +
        '<div class="action-links">' +
        '<a href="history.html">📊 View Progress</a>' +
        '<a href="mistakes.html">📖 Review Mistakes</a>' +
        '</div>' +
        explanationsHtml;

    document.getElementById('results').innerHTML = resultsHtml;
    document.getElementById('results').classList.add('show');
    document.getElementById('submit-btn').disabled = true;
    document.getElementById('submit-btn').textContent = 'SUBMITTED';
    document.getElementById('results').scrollIntoView({behavior: 'smooth'});

    sessionStorage.setItem('vocab_submitted_state_' + CURRENT_ID, JSON.stringify({
        formHTML: document.getElementById('quiz-form').innerHTML,
        resultsHTML: document.getElementById('results').innerHTML,
        timer: document.getElementById('timer-display').textContent,
        progress: document.getElementById('progress-count').textContent
    }));
}

function nextSet() {
    const done = JSON.parse(localStorage.getItem('done_set_ids') || '[]');
    if (CURRENT && !done.includes(CURRENT.id)) {
        done.push(CURRENT.id);
        localStorage.setItem('done_set_ids', JSON.stringify(done));
    }
    sessionStorage.removeItem('current_set_id');
    sessionStorage.removeItem('vocab_submitted_state_' + CURRENT_ID);
    location.reload();
}

function showAllDone() {
    document.body.innerHTML = `
        <div class="all-done">
            <h1>🎉 All Quiz Sets Completed!</h1>
            <p style="color:#666;font-size:17px;margin:24px 0;">
                You've finished all ${ALL_SETS.length} quiz sets.<br>
                Great work!
            </p>
            <div class="actions">
                <button onclick="resetAndContinue()">🔄 Reset and Practice Again</button>
                <button class="secondary" onclick="showGenerateMore()">📚 Generate More</button>
            </div>
            <div style="margin-top:40px;font-size:13px;color:#888;">
                <a href="history.html" style="color:#1a4d8f;">📊 History</a> &nbsp;|&nbsp;
                <a href="mistakes.html" style="color:#1a4d8f;">📖 Mistakes Book</a>
            </div>
        </div>`;
}

function resetAndContinue() {
    if (confirm('Reset all "done" records? Your history and mistakes are NOT affected.')) {
        localStorage.removeItem('done_set_ids');
        location.reload();
    }
}

function showGenerateMore() {
    alert("To generate more quiz sets, run on your computer:\\n\\ncd ~/code/isee-vocab\\nuv run main.py 50\\n\\nThen: git add . && git commit -m 'More sets' && git push");
}

function showError(msg) {
    document.getElementById('start-card').innerHTML = `
        <h2 style="color:#c62828;">⚠️ Error</h2>
        <p style="color:#555;font-family:Arial;font-size:14px;">${msg}</p>`;
}

init();
</script>

</body>
</html>'''


def render_history_html():
    return '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vocab Progress History</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
* { box-sizing: border-box; }
body { font-family: "Georgia", serif; max-width: 860px; margin: 0 auto; padding: 24px 20px 60px; background: #fafaf7; }
header { border-bottom: 3px double #333; padding-bottom: 16px; margin-bottom: 28px; }
h1 { font-size: 24px; margin: 0 0 6px; }
h2 { font-size: 18px; margin: 32px 0 16px; padding-left: 12px; border-left: 4px solid #1a4d8f; }
.subtitle { font-size: 13px; color: #555; font-style: italic; }
.back-link { display: inline-block; margin-bottom: 20px; color: #1a4d8f; text-decoration: none; font-family: Arial, sans-serif; font-size: 14px; }
.stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; }
.stat-card { background: white; padding: 18px 16px; border: 1px solid #ddd; border-radius: 6px; text-align: center; font-family: Arial, sans-serif; }
.stat-value { font-size: 28px; font-weight: bold; color: #1a4d8f; }
.stat-label { font-size: 11px; color: #666; text-transform: uppercase; letter-spacing: 1px; margin-top: 6px; }
.chart-section { background: white; padding: 20px; border: 1px solid #ddd; border-radius: 6px; margin-top: 12px; }
.chart-container { position: relative; height: 320px; }
table { width: 100%; border-collapse: collapse; background: white; font-family: Arial, sans-serif; font-size: 13px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); border-radius: 6px; overflow: hidden; }
th { background: #1a4d8f; color: white; padding: 10px 12px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; }
td { padding: 10px 12px; border-bottom: 1px solid #eee; }
tr:hover td { background: #fafafa; }
button { background: #1a4d8f; color: white; border: none; padding: 8px 18px; font-size: 13px; border-radius: 4px; cursor: pointer; font-family: Arial, sans-serif; }
button.danger { background: #c62828; }
.clear-section { margin-top: 20px; text-align: right; }
.empty-state { text-align: center; padding: 60px 20px; color: #888; }
@media (max-width: 600px) { .stats-grid { grid-template-columns: repeat(2, 1fr); } table { font-size: 11px; } th, td { padding: 8px 6px; } }
</style>
</head>
<body>
<a href="index.html" class="back-link">← Back to Practice</a>
<header><h1>VOCAB PROGRESS</h1><div class="subtitle">Practice history</div></header>
<div id="content"></div>
<script>
function render() {
    const sessions = JSON.parse(localStorage.getItem('vocab_sessions') || '[]');
    const content = document.getElementById('content');
    if (sessions.length === 0) {
        content.innerHTML = '<div class="empty-state"><h2 style="border:none;padding:0;">📊 No sessions yet</h2><p>Complete a quiz to start tracking.</p></div>';
        return;
    }
    const totalQ = sessions.reduce((s, x) => s + x.total, 0);
    const totalCorrect = sessions.reduce((s, x) => s + x.correct, 0);
    const avgAcc = Math.round(totalCorrect / totalQ * 100);
    let trend = '—', trendColor = '#888';
    if (sessions.length >= 4) {
        const recent = sessions.slice(-3).reduce((s, x) => s + x.accuracy, 0) / 3;
        const earlier = sessions.slice(0, 3).reduce((s, x) => s + x.accuracy, 0) / 3;
        const diff = Math.round(recent - earlier);
        if (diff > 0) { trend = '↑ +' + diff + '%'; trendColor = '#27ae60'; }
        else if (diff < 0) { trend = '↓ ' + diff + '%'; trendColor = '#c62828'; }
        else trend = '→ 0%';
    }
    let html = '<div class="stats-grid">' +
        '<div class="stat-card"><div class="stat-value">' + sessions.length + '</div><div class="stat-label">Quizzes</div></div>' +
        '<div class="stat-card"><div class="stat-value">' + totalQ + '</div><div class="stat-label">Questions</div></div>' +
        '<div class="stat-card"><div class="stat-value">' + avgAcc + '%</div><div class="stat-label">Avg Accuracy</div></div>' +
        '<div class="stat-card"><div class="stat-value" style="color:' + trendColor + '">' + trend + '</div><div class="stat-label">Trend</div></div>' +
        '</div>' +
        '<h2>📈 Accuracy Over Time</h2>' +
        '<div class="chart-section"><div class="chart-container"><canvas id="acc-chart"></canvas></div></div>' +
        '<h2>📋 All Sessions</h2>' +
        '<table><thead><tr><th>Date</th><th>Set #</th><th>Score</th><th>Acc.</th><th>Time</th></tr></thead><tbody>';
    sessions.slice().reverse().forEach(s => {
        const time = new Date(s.timestamp).toLocaleTimeString('en-US', {hour:'2-digit',minute:'2-digit'});
        const dur = Math.floor(s.duration_sec/60) + 'm' + (s.duration_sec%60) + 's';
        const accColor = s.accuracy >= 80 ? '#27ae60' : (s.accuracy >= 60 ? '#f57c00' : '#c62828');
        html += '<tr><td>' + s.date + '<br><span style="color:#999;font-size:10px;">' + time + '</span></td>' +
            '<td>#' + (s.set_number || '?') + '</td>' +
            '<td>' + s.correct + '/' + s.total + '</td>' +
            '<td style="color:' + accColor + ';font-weight:bold;">' + s.accuracy + '%</td>' +
            '<td>' + dur + '</td></tr>';
    });
    html += '</tbody></table><div class="clear-section"><button class="danger" onclick="if(confirm(\\'Clear all?\\')){localStorage.removeItem(\\'vocab_sessions\\');render();}">Clear All History</button></div>';
    content.innerHTML = html;
    const ctx = document.getElementById('acc-chart');
    new Chart(ctx, {
        type: 'line',
        data: { labels: sessions.map(s => s.date), datasets: [{ label: 'Accuracy', data: sessions.map(s => s.accuracy), borderColor: '#1a4d8f', backgroundColor: 'rgba(26,77,143,0.1)', borderWidth: 2.5, pointRadius: 5, tension: 0.3, fill: true }] },
        options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, max: 100, ticks: { callback: v => v + '%' } } } }
    });
}
render();
</script>
</body>
</html>'''


def render_mistakes_html():
    return '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mistakes Book</title>
<style>
* { box-sizing: border-box; }
body { font-family: "Georgia", serif; max-width: 760px; margin: 0 auto; padding: 24px 20px 60px; background: #fafaf7; }
header { border-bottom: 3px double #333; padding-bottom: 16px; margin-bottom: 28px; }
h1 { font-size: 24px; margin: 0 0 6px; }
.subtitle { font-size: 13px; color: #555; font-style: italic; }
.back-link { display: inline-block; margin-bottom: 20px; color: #1a4d8f; text-decoration: none; font-family: Arial, sans-serif; font-size: 14px; }
.stats { background: white; padding: 14px 18px; border: 1px solid #ddd; border-radius: 4px; margin-bottom: 24px; font-family: Arial, sans-serif; font-size: 14px; }
.mistake-card { background: white; padding: 14px 18px; margin-bottom: 10px; border-left: 4px solid #c62828; border-radius: 0 4px 4px 0; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
.mistake-word { font-size: 18px; font-weight: bold; color: #1a4d8f; }
.mistake-stem { color: #333; font-size: 14px; margin-top: 4px; }
.mistake-answer { font-size: 13px; margin: 6px 0; }
.mistake-answer .user { color: #c62828; }
.mistake-answer .correct { color: #2e7d32; }
.mistake-exp { font-size: 12px; color: #666; margin-top: 6px; font-style: italic; }
.empty-state { text-align: center; padding: 60px 20px; color: #888; }
button { background: #c62828; color: white; border: none; padding: 8px 18px; font-size: 13px; border-radius: 4px; cursor: pointer; font-family: Arial, sans-serif; }
.clear-section { margin-top: 20px; text-align: right; }
</style>
</head>
<body>
<a href="index.html" class="back-link">← Back to Practice</a>
<header><h1>MISTAKES BOOK</h1><div class="subtitle">Words you missed — review these</div></header>
<div id="content"></div>
<script>
function render() {
    const mistakes = JSON.parse(localStorage.getItem('vocab_mistakes') || '[]');
    const content = document.getElementById('content');
    if (mistakes.length === 0) {
        content.innerHTML = '<div class="empty-state"><p>🎯 No mistakes yet! Keep practicing.</p></div>';
        return;
    }
    let html = '<div class="stats"><strong>' + mistakes.length + '</strong> words in your mistakes book</div>';
    mistakes.slice().reverse().forEach(m => {
        html += '<div class="mistake-card">' +
            '<div class="mistake-word">' + (m.word || '(word)') + ' <span style="font-size:11px;color:#999;font-weight:normal;font-family:Arial;">(' + m.type + ')</span></div>' +
            '<div class="mistake-stem">' + m.stem + '</div>' +
            '<div class="mistake-answer">You: <span class="user">' + m.user + '</span> → Correct: <span class="correct">' + m.correct + '</span></div>' +
            '<div class="mistake-exp">' + m.explanation + '</div>' +
            '</div>';
    });
    html += '<div class="clear-section"><button onclick="if(confirm(\\'Clear all mistakes?\\')){localStorage.removeItem(\\'vocab_mistakes\\');render();}">Clear All</button></div>';
    content.innerHTML = html;
}
render();
</script>
</body>
</html>'''


# ============ Entry Point ============
if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1].lower()
        if arg == "reset":
            confirm = input("⚠️  Delete quiz-sets.json? (yes/no): ")
            if confirm.lower() == "yes":
                if QUIZ_SETS_FILE.exists():
                    QUIZ_SETS_FILE.unlink()
                    print("✅ quiz-sets.json deleted.")
                else:
                    print("File doesn't exist.")
            else:
                print("Cancelled.")
            sys.exit(0)
        try:
            batch_size = int(arg)
        except ValueError:
            print(f"⚠️  Invalid argument. Use a number or 'reset'.")
            sys.exit(1)
    else:
        batch_size = DEFAULT_BATCH_SIZE

    sets = run_batch(batch_size)

    with open(INDEX_FILE, "w", encoding="utf-8") as f:
        f.write(render_index_html())
    with open(HISTORY_PAGE, "w", encoding="utf-8") as f:
        f.write(render_history_html())
    with open(MISTAKES_FILE, "w", encoding="utf-8") as f:
        f.write(render_mistakes_html())

    print(f"📄 HTML files generated:")
    print(f"   - {INDEX_FILE}")
    print(f"   - {HISTORY_PAGE}")
    print(f"   - {MISTAKES_FILE}")
    print(f"   - {QUIZ_SETS_FILE}  ({len(sets)} sets)")
    print(f"\n✅ Done!")